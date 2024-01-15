##############################################################################################################################################################
##  Script that generates a report for all properties                                                                                                       ##
##                                                                                                                                                          ##
##  Made by : Sid                                                                                                                                           ##    
##                                                                                                                                                          ##
##  Get all items from Monday and keys from Keynest                                                                                                         ##
##                                                                                                                                                          ##
##  For each item in Monday :                                                                                                                               ##
##      1) Is the property available                                                                                                                        ##
##      2) When is the next checkout ?                                                                                                                      ##
##      3) When is the next checkin                                                                                                                         ##
##                                                                                                                                                          ##
##############################################################################################################################################################


# Libraries
from graphql import graphql_query
from queries import boards_pagination, get_boards_graphql, get_item_cursor, items_pagination, new_get_item_query, new_item_pagination, new_paginated_query, paginated_query, parse_board_data, get_item_query, parse_item_data, change_item_value, move_item_to_group, create_group
from json import loads
from datetime import datetime, timedelta
import calendar
import smtplib
import os
from dotenv import load_dotenv
import csv
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import COMMASPACE
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from api import APIEndPoint
from datetime import date
import pandas as pd
from openpyxl import load_workbook

# Script
load_dotenv()
now_date = datetime.now()
keys=[]
check_in_list=[]
check_out_list=[]
formated_items=[]
same_day_list=list()
action=list()
action_list=[]

filtre=['Manual Reservation Processes', 'Automated Reservation Processes', 'Bookings Audit 2022', 'Booking Partner Board', 'Revenue Management Board', 'Bookings Audit 2021', 'Manual Reservation Processes', 'Bookings audit 2020']
list=[]
i=0
today_date = date.today()
today = today_date.strftime("%d/%m/%Y")

def get_next_month():
    # Get the current date
    current_date = datetime.now()

    # Calculate the next month
    next_month = current_date.replace(day=1) + timedelta(days=32)

    # Format the next month as a string
    next_month_str = next_month.strftime("%B")

    return next_month_str

def get_current_month():

    # Get the current date
    current_date = datetime.now()

    # Calculate the next month
    current_month = current_date.replace(day=1)

    # Format the next month as a string
    current_month_str = current_month.strftime("%B")

    return current_month_str


def get_formated_items():
    # Define current and next month
    next_month=get_next_month()
    current_month=get_current_month()
    INCLUDE_GROUPS = os.getenv("INCLUDE_GROUPS").split(",") if os.getenv("INCLUDE_GROUPS") != "" else [] 
    EXCLUDE_GROUPS = os.getenv("EXCLUDE_GROUPS").split(",") if os.getenv("EXCLUDE_GROUPS") != "" else [] 
    INCLUDE_GROUPS.append(current_month)
    INCLUDE_GROUPS.append(next_month)
    boards = parse_board_data(paginated_query(get_boards_graphql(), boards_pagination))
    for board in boards:
        if board['name'] not in filtre:
            for group in board["groups"]:
                if len(INCLUDE_GROUPS) != 0 and group["title"] not in INCLUDE_GROUPS:
                    continue

                if len(EXCLUDE_GROUPS) != 0 and group["title"] in EXCLUDE_GROUPS:
                    continue

                items = parse_item_data(new_paginated_query(new_get_item_query(board["id"], group["id"]), new_item_pagination, get_item_cursor))

                for item in items:
                    try:
                        # Guest Name, Booking Dates, Property Allocated, 
                        # Property Booked, Checkin Time, Checkout Time

                        # Get Columns
                        Status = [x for x in item['column_values'] if x['column']['title'] == 'Status']
                        Timeline = [x for x in item['column_values'] if x['column']['title'] == 'Timeline']
                        Check_in_time = [x for x in item['column_values'] if x['column']['title'] == 'Check-in Time']
                        Check_out_time = [x for x in item['column_values'] if x['column']['title'] == 'Check-out Time']
                        allocated_property = [x for x in item['column_values'] if x['column']['title'] == 'Allocated Property']
                        flat_booked = [x for x in item['column_values'] if x['column']['title'] == 'Flat booked']
                        form_completed=[x for x in item['column_values'] if x['column']['title'] == 'Check-in form status']

                        # Check if columns exist
                        if len(Status) == 0:
                            raise Exception(f"No Status for {item['name']}")
                        if len(Timeline) == 0:
                            raise Exception(f"No Timeline for {item['name']}")
                        if len(Check_in_time) == 0:
                            raise Exception(f"No Check-in Time for {item['name']}")
                        if len(Check_out_time) == 0:
                            raise Exception(f"No Check-out Time for {item['name']}")
                        if len(allocated_property) == 0:
                            raise Exception(f"No allocated property for {item['name']}")
                        if len(flat_booked) == 0:
                            raise Exception(f"No flat booked for {item['name']}")
                        if len(form_completed) == 0:
                            raise Exception(f"No Check-in form for {item['name']}")

                        # Get correct variables
                        Status=Status[0]['text']
                        check_in_date=Timeline[0]['text'].split(' - ')[0]
                        check_out_date=Timeline[0]['text'].split(' - ')[1]
                        allocated_property = allocated_property[0]['text']
                        flat_booked = flat_booked[0]['text']
                        Check_in_time = Check_in_time[0]['text']
                        Check_out_time = Check_out_time[0]['text']
                        form_completed=form_completed[0]['text']


                        # Get tag
                        tag = flat_booked.split(' - ')[0]
                        
                        # Convert the date format
                        date_obj = datetime.strptime(check_in_date, "%Y-%m-%d")
                        converted_date_1 = date_obj.strftime("%d/%m/%Y")      
                        
                        date_obj = datetime.strptime(check_out_date, "%Y-%m-%d")
                        converted_date_2 = date_obj.strftime("%d/%m/%Y") 

                        # New format
                        new_item = {'guest': item['name'], 'status': Status ,'tag': tag , 'allocated_property':allocated_property, 'flat_booked': flat_booked, 'checkin_date' : converted_date_1, 'checkout_date' : converted_date_2, 'checkin_time': Check_in_time, 'checkout_time' : Check_out_time, 'form': form_completed , 'group': group['title']}
                        formated_items.append(new_item)
                    
                    except Exception as e:
                        print(e)

    return formated_items

def set_column_width(filename, column_widths):
    # Load the workbook
    wb = load_workbook(filename)

    # Select the active sheet
    sheet = wb.active

    # Set the width of each column
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    # Save the workbook
    wb.save(filename)

def send_email(property_status):

    export_to_excel(property_status)
    # Set the width of the columns
    column_widths = {'A': 25, 'B': 13, 'C': 13, 'D': 13, 'E': 13, 'F': 13}
    set_column_width('properties_report.xlsx', column_widths)

    # The Excel file
    filename = "properties_report.xlsx"
    attachment_path = os.path.abspath(filename)

    # Email details
    sender_email = 'smtp@theflexliving.com'
    receiver_emails = ['sid@theflexliving.com']#, 'michael@theflexliving.com']
    subject = "[AUTOMATION] - Properties report"
    body = "Please find attached the Excel file containing the list of properties with their status"

    # Create message container
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ", ".join(receiver_emails)
    msg['Subject'] = subject

    # Add body to email
    msg.attach(MIMEText(body, "plain"))

    # Attach the Excel file
    with open(attachment_path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read()) 
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {filename}",
        )
        msg.attach(part)

    # Convert message object to string
    text = msg.as_string()

    # Log in to SMTP server and send email
    try:
        smtp_server = smtplib.SMTP('smtp.gmail.com', 587)
        smtp_server.starttls()
        smtp_server.login(sender_email, "")
        smtp_server.sendmail(sender_email, receiver_emails, text)
        smtp_server.quit()
    
    except Exception as e:
        print("An error occurred while sending the email:", e)

def export_to_excel(property_status):
    data = []
    for property, status in property_status.items():
        current_checkin = status['current_checkin'].strftime('%d/%m/%Y') if status['current_checkin'] is not None else None
        current_checkout = status['current_checkout'].strftime('%d/%m/%Y') if status['current_checkout'] is not None else None
        next_checkin = status['next_checkin'] if status['next_checkin'] is not None else "N/A"
        next_checkout = status['next_checkout']if status['next_checkout'] is not None else "N/A"

        # If both next checkin and checkout are None, set availability to "Available"
        if (status['is_available']==True) :
            availability = "Available"
        else:
            availability = "Not Available" if status['is_available'] is False else "Available"

        data.append([property, availability, current_checkin, current_checkout, next_checkin, next_checkout])

    df = pd.DataFrame(data, columns=['Property', 'Availability', 'Current Check-in', 'Current Check-out', 'Next Check-in', 'Next Check-out'])

    # Write the DataFrame to an Excel file
    df.to_excel('properties_report.xlsx', index=False)

def find_next_booking(item, booked_flats):
    item_checkout_date = datetime.strptime(item['checkout_date'], '%d/%m/%Y').date()
    next_flat = None
    min_time_diff = float('inf')

    for flat in booked_flats:
        checkin_date = datetime.strptime(flat['checkin_date'], '%d/%m/%Y').date()
        checkout_date = datetime.strptime(item['checkout_date'], '%d/%m/%Y').date()
        time_diff = (checkin_date - item_checkout_date).days

        # Check for bookings in the future (including the same day as the current checkout)
        if time_diff >= 0 and time_diff < min_time_diff:
            next_flat = flat
            min_time_diff = time_diff

    return next_flat

def check_availability(formated_items):
    """
    Check availability of properties based on formated items data.
    """
    today = datetime.now().date()
    property_bookings = {}

    filtered_properties = [prop for prop in formated_items if prop['allocated_property'] != None and prop['allocated_property'] != "" ]

    # Get all bookings for each property sorted by check-in date
    for item in filtered_properties:
        property_name = item['allocated_property']
        checkin_date = datetime.strptime(item['checkin_date'], '%d/%m/%Y').date()
        checkout_date = datetime.strptime(item['checkout_date'], '%d/%m/%Y').date()
        booked_flats = [prop for prop in filtered_properties if prop['allocated_property'] == item['allocated_property']]
        next_res = find_next_booking(item, booked_flats)

        if property_name not in property_bookings:
            property_bookings[property_name] = []
        
        property_bookings[property_name].append((checkin_date, checkout_date, item['guest'], next_res['checkin_date'] if next_res else None, next_res['checkout_date'] if next_res else None))

    property_availability = {}

    # Calculate current and next availability
    for property_name, bookings in property_bookings.items():
        # sort bookings by checkin_date
        bookings.sort(key=lambda x: x[0])

        current_checkin = current_checkout = next_checkin = next_checkout = None
        is_available = True  # Property is initially available

        for i, booking in enumerate(bookings):
            checkin_date, checkout_date, guest, next_checkin_date, next_checkout_date = booking
            
            # If the property is currently booked
            if checkin_date <= today <= checkout_date:
                current_checkin = checkin_date
                current_checkout = checkout_date
                is_available = False  # Property is not available

            # If the property has a future booking
            elif today < checkin_date:
                next_checkin = checkin_date
                next_checkout = checkout_date
                is_available = False
                break

        property_availability[property_name] = {
            'current_checkin': current_checkin,
            'current_checkout': current_checkout,
            'next_checkin': next_checkin,
            'next_checkout': next_checkout,
            'is_available': is_available
        }

    return property_availability


def main():
    """
    The main function to execute the program.
    """

    # Fetch and format the data from the source
    formatted_items = get_formated_items()
    
    # For tests
    #filtered_properties = [prop for prop in formated_items if prop['allocated_property'] == "2B N7 - 138 Xchange Point"]
    
    # Check availability of each property
    property_availability = check_availability(formatted_items)

    # Export the availability data to an Excel file
    export_to_excel(property_availability)

    send_email(property_availability)

    print("*** SCRIPT ENDED WITH NO ERROR *** ")
    

# Run the main function
if __name__ == "__main__":
    main()
