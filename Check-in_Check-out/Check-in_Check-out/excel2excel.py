import pandas as pd
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


def set_column_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 0.5) * 1.1
        sheet.column_dimensions[column_letter].width = adjusted_width


def formatExcel():
    # Create a new workbook
    workbook = Workbook()

    # Select the active worksheet
    worksheet = workbook.active

    # list of head row indexes of each dataframe
    head_ind = [1]

    # list of checkin actions to filter
    checkin_actions = ['RELOCATE', 'TO BE CANCELLED', 'SEND CHECKIN', 'DELAYED CHECKIN',
                       'CHECK-IN FORM NOT COMPLETED - SEND REMINDER',
                       'URGENT - CLASH - Ask for delayed Checking',
                       'SEND CHECKIN INSTRUCTIONS - CHECK-IN FORM COMPLETED']
    
    keys_actions =['NO SHOW', 'KEYS NOT RETURNED', 'KEYS RETURNED', 'NO KEYS FOUND']

    # read check_in_check_out.xlsx into dataframe
    df = pd.read_excel('organized_data.xlsx')

    # preprocess the dataframe
    df['action'] = df['action'].str.strip()
    if df['keys'].isna().any():
        # Replace NaN values in the 'keys' column with an empty string
        df['keys'] = df['keys'].fillna("")
    else :
        df['keys'] = df['keys'].str.strip()

    df['group'] = df['group'].str.strip()
    df['action'] = df['action'].replace('SEND CHECKOUT INSTRUCTIONS', 'SEND CHECKOUT')

    # dataframe for errors
    errors_df = df[ (df['action'] == 'URGENT - CLASH' )|
                     (df['action'] == 'Property not allocated') ]
    errors_str = errors_df.to_csv(index=False)

    head_ind.append(len(errors_df) + 4)
    # dataframe for checkins
    # checkins_df = df[df['action'].isin(checkin_actions)]
    checkins_df = df[(df['action'].isin(checkin_actions)) |
                     ((df['action'] == 'URGENT - Early check-in & Late check-out') &
                      (df['group'] == '1 Week Before Check In'))]
    checkins_str = checkins_df.to_csv(index=False)

    head_ind.append(head_ind[-1] + len(checkins_df) + 5)
   
    # dataframe for checkouts
    # checkouts_df = df[df['action'] == 'SEND CHECKOUT']
    checkouts_df = df[((df['action'] == 'SEND CHECKOUT')) |
                      ((df['action'] == 'NO CLASH') & (df['group'] == 'Check Outs')  )]
    checkouts_str = checkouts_df.to_csv(index=False)
    
    
    head_ind.append(head_ind[-1] + len(checkouts_df) + 5)

    # dataframe for keynest
    keynest_df = df[(df['keys'].isin(keys_actions))]
    keynest_str = keynest_df.to_csv(index=False)

    # strings of csv format for adding lines between dataframes
    head1_str = "ERRORS,,,,,,,,,"
    head2_str = ",,,,,,,,,\nCheckins,,,,,,,,,"
    head3_str = ",,,,,,,,,\n,,,,,,,,,\n,,,,,,,,,\nCheckouts,,,,,,,,,"

    # combine all the strings of csv format into one csv-format string
    combined_str = head1_str + '\n' + errors_str + head2_str \
                   + '\n' + checkins_str + head3_str + '\n' \
                   + checkouts_str

    # Add the "Keynest" header at the end
    combined_str += ",,,,,,,,,\n,,,,,,,,,\n,,,,,,,,,\nKeynest,,,,,,,,,"+ '\n' \
                   + keynest_str

    # Create a CSV reader object from the string
    csv_reader = csv.reader(combined_str.splitlines())

    # Iterate over each row in the CSV reader
    for row_index, row in enumerate(csv_reader, start=1):
        # Write each value to the corresponding cell in the worksheet
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index).value = value

    # format the workbook with colors
    bg_color = '164F4C'
    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

    for row_ind in head_ind:
        for col_ind in range(1, 12):
            cell = worksheet.cell(row=row_ind, column=col_ind)
            cell.fill = fill
            cell.font = Font(bold=True, color='FFFFFF')

    set_column_width(worksheet)

    # Save the workbook
    workbook.save('output.xlsx')
    workbook.close()


formatExcel()