##############################################################################################################################################################
##  Script that generates a report for checkins & checkouts                                                                                                 ##
##                                                                                                                                                          ##
##  Made by : Sid                                                                                                                                           ##    
##                                                                                                                                                          ##
##  Get all items from Monday and keys from Keynest                                                                                                         ##
##                                                                                                                                                          ##
##  For each item in Monday :                                                                                                                               ##
##      1) Filter items : Ignore empty allocated properties / empty flat booked / Cancelled reservations / to be cancelled items                            ##
##      2) Check if reservation is a checkin or a checkout                                                                                                  ##
##      3)           
## 
## 
## 
## 
##############################################################################################################################################################

# Libraries
from graphql import graphql_query
from queries import boards_pagination, get_boards_graphql, get_item_cursor, items_pagination, new_get_item_query, new_item_pagination, new_paginated_query, paginated_query, parse_board_data, get_item_query, parse_item_data, change_item_value, move_item_to_group, create_group, new_get_item_by_board_id_query
from json import loads
from datetime import datetime, timedelta
import calendar
import smtplib
import os
from dotenv import load_dotenv
import csv
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import COMMASPACE
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from excel2excel import formatExcel
from api import APIEndPoint
from datetime import date

# Script
load_dotenv()
now_date = datetime.now()
keys=[]
check_in_list=[]
check_out_list=[]
formated_items=[]
same_day_list=list()
action=list()
action_list=[]

filtre=['Manual Reservation Processes', 'Automated Reservation Processes', 'Bookings Audit 2022', 'Booking Partner Board', 'Revenue Management Board', 'Bookings Audit 2021', 'Manual Reservation Processes', 'Bookings audit 2020']
list=[]
i=0

class KeynestClient():

    def __init__(self):
        # initiate keynest client
        self.keynest_client = self._initiate_client()
        
    def _initiate_client(self):
        return APIEndPoint('')

def no_show_keys(item, key_list):
    
    filtered_keys= [prop for prop in key_list if prop['KeyName'] == item['allocated_property']]

    if len(filtered_keys) > 0:
        
        # list of keys not dropped 
        not_dropped=[]
        no_show=[]
        for key in filtered_keys:

            # Conditions for checkints : Checkin was yesterday but key was not collected
            if key['StatusType'] == "In Store" and item['group'] !='Check Outs':
                no_show.append(key)

            if len(no_show) == len(filtered_keys):
                action= item
                action.update({"keys": "NO SHOW"})

def keys_dropped(item, key_list):

    filtered_keys= [prop for prop in key_list if prop['KeyName'] == item['allocated_property']]

    if len(filtered_keys) > 0:
        
        # list of keys not dropped 
        not_dropped=[]
        for key in filtered_keys:
            # Variables
            """
            if key['LastMovement'] != None:
                datetime_obj = datetime.strptime(key['LastMovement'], "%Y-%m-%dT%H:%M:%S.%f")
                date = datetime_obj.strftime("%d/%m/%Y")
                time = datetime_obj.strftime("%H:%M")
                today = date.today()
                print(today)
            """
            # Conditions for checkouts : Checkout is today but key was not dropped to Keynest
            if key['StatusType'] == "In Use":
                not_dropped.append(key)

            if len(not_dropped)>0:
                action= item
                action.update({"keys": "KEYS NOT RETURNED"})
            else : 
                action=item
                action.update({"keys": "KEYS RETURNED"})
    
    else :
        action=item
        action.update({"keys": "NO KEYS FOUND"})
    
    return action

                  
                      
def compare_dates_check_out(item, date_param, keys_list):
    
    send_checkout=False
    # Obtenir la date d'aujourd'hui
    today = datetime.now().date()

    # Convertir la date donnée en paramètre en objet de type 'date'
    date = datetime.strptime(date_param, "%d/%m/%Y").date()

    # Calculer la différence entre les deux dates
    difference = (date - today).days

    # Vérifier si la différence est comprise entre 0 et 3 jours
    if 0 <= difference <= 3:
        #print(f"Check-out - {item['guest']}")
        send_checkout=True
        check_out_list.append({'guest':item["guest"], 'check_out_date': date_param})
        #send_email()  # Appeler la fonction send_email() si la condition est remplie
    
    if difference == 0 :
        keys_dropped(item, keys_list)
        
    return send_checkout

def compare_dates_check_in(item, date_param, keys_list):
    send_checkin=False
    # Obtenir la date d'aujourd'hui
    today = datetime.now().date()

    # Convertir la date donnée en paramètre en objet de type 'date'
    date = datetime.strptime(date_param, "%d/%m/%Y").date()

    # Calculer la différence entre les deux dates
    difference = (date - today).days

    # Vérifier si la différence est comprise entre 0 et 3 jours
    if 0 <= difference <= 7:
        #print(f"Check-in - {item['guest']}")
        send_checkin= True
        check_in_list.append({'guest':item["guest"], 'check_in_date': date_param})
        #send_email()  # Appeler la fonction send_email() si la condition est remplie       

    if difference == -1:
        keys_dropped(item, keys_list)
    
    return send_checkin 

def get_formated_items():
    INCLUDE_GROUPS = os.getenv("INCLUDE_GROUPS").split(",") if os.getenv("INCLUDE_GROUPS") != "" else [] 
    EXCLUDE_GROUPS = os.getenv("EXCLUDE_GROUPS").split(",") if os.getenv("EXCLUDE_GROUPS") != "" else [] 
    boards = parse_board_data(paginated_query(get_boards_graphql(), boards_pagination))
    for board in boards:
        if board['name'] not in filtre:
            for group in board["groups"]:
                if len(INCLUDE_GROUPS) != 0 and group["title"] not in INCLUDE_GROUPS:
                    continue

                if len(EXCLUDE_GROUPS) != 0 and group["title"] in EXCLUDE_GROUPS:
                    continue

                items = parse_item_data(new_paginated_query(new_get_item_query(board["id"], group["id"]), new_item_pagination, get_item_cursor))

                for item in items:
                    try:
                        # Guest Name, Booking Dates, Property Allocated, 
                        # Property Booked, Checkin Time, Checkout Time

                        # Get Columns
                        Status = [x for x in item['column_values'] if x['column']['title'] == 'Status']
                        Timeline = [x for x in item['column_values'] if x['column']['title'] == 'Timeline']
                        Check_in_time = [x for x in item['column_values'] if x['column']['title'] == 'Check-in Time']
                        Check_out_time = [x for x in item['column_values'] if x['column']['title'] == 'Check-out Time']
                        allocated_property = [x for x in item['column_values'] if x['column']['title'] == 'Allocated Property']
                        flat_booked = [x for x in item['column_values'] if x['column']['title'] == 'Flat booked']
                        form_completed=[x for x in item['column_values'] if x['column']['title'] == 'Check-in form status']

                        # Check if columns exist
                        if len(Status) == 0:
                            raise Exception(f"No Status for {item['name']}")
                        if len(Timeline) == 0:
                            raise Exception(f"No Timeline for {item['name']}")
                        if len(Check_in_time) == 0:
                            raise Exception(f"No Check-in Time for {item['name']}")
                        if len(Check_out_time) == 0:
                            raise Exception(f"No Check-out Time for {item['name']}")
                        if len(allocated_property) == 0:
                            raise Exception(f"No allocated property for {item['name']}")
                        if len(flat_booked) == 0:
                            raise Exception(f"No flat booked for {item['name']}")
                        if len(form_completed) == 0:
                            raise Exception(f"No Check-in form for {item['name']}")

                        # Get correct variables
                        Status=Status[0]['text']
                        check_in_date=Timeline[0]['text'].split(' - ')[0]
                        check_out_date=Timeline[0]['text'].split(' - ')[1]
                        allocated_property = allocated_property[0]['text']
                        flat_booked = flat_booked[0]['text']
                        Check_in_time = Check_in_time[0]['text']
                        Check_out_time = Check_out_time[0]['text']
                        form_completed=form_completed[0]['text']


                        # Get tag
                        tag = flat_booked.split(' - ')[0]
                        
                        # Convert the date format
                        date_obj = datetime.strptime(check_in_date, "%Y-%m-%d")
                        converted_date_1 = date_obj.strftime("%d/%m/%Y")      
                        
                        date_obj = datetime.strptime(check_out_date, "%Y-%m-%d")
                        converted_date_2 = date_obj.strftime("%d/%m/%Y") 

                        # New format
                        new_item = {'guest': item['name'], 'status': Status ,'tag': tag , 'allocated_property':allocated_property, 'flat_booked': flat_booked, 'checkin_date' : converted_date_1, 'checkout_date' : converted_date_2, 'checkin_time': Check_in_time, 'checkout_time' : Check_out_time, 'form': form_completed , 'group': group['title']}
                        formated_items.append(new_item)
                    
                    except Exception as e:
                        print(e)

    return formated_items

def organize_data(action_list):
    # Create Excel workbook
    workbook = openpyxl.Workbook()

    # Create worksheet
    sheet = workbook.active
    sheet.title = "Data"

    # Define column names
    column_names = ['guest', 'tag', 'allocated_property', 'flat_booked', 'checkin_date', 'checkout_date', 'checkin_time', 'checkout_time', 'group', 'action']

    # Add column names to the worksheet
    sheet.append(column_names)

    # Sort the action_list based on the "action" column
    sorted_action_list = sorted(action_list, key=lambda x: x['action'])

    # Set the starting row index for each section
    checkin_row = 2
    checkout_row = 2
    errors_row = 2

    # Iterate over the sorted action_list
    for guest in sorted_action_list:
        # Check the value of the "action" column
        if guest['action'] == 'SEND CHECKIN INSTRUCTIONS':
            # Add the "Checkins" header and column names
            if checkin_row == 2:
                sheet.cell(row=1, column=1).font = Font(bold=True)
                sheet.cell(row=1, column=1, value='Checkins')
                for col_num, column_name in enumerate(column_names, start=2):
                    sheet.cell(row=1, column=col_num).font = Font(bold=True)
                    sheet.cell(row=1, column=col_num, value=column_name)
            # Add the data for "SEND CHECKIN INSTRUCTIONS" rows
            for col_num, column_name in enumerate(column_names, start=1):
                sheet.cell(row=checkin_row, column=col_num, value=guest[column_name])
            checkin_row += 1
        elif guest['action'] in ('SEND CHECKOUT INSTRUCTIONS', 'RELOCATE'):
            # Add the "Checkouts" header and column names
            if checkout_row == 2:
                sheet.cell(row=1, column=1).font = Font(bold=True)
                sheet.cell(row=1, column=1, value='Checkouts')
                for col_num, column_name in enumerate(column_names, start=2):
                    sheet.cell(row=1, column=col_num).font = Font(bold=True)
                    sheet.cell(row=1, column=col_num, value=column_name)
            # Add the data for "SEND CHECKOUT INSTRUCTIONS" or "RELOCATE" rows
            for col_num, column_name in enumerate(column_names, start=1):
                sheet.cell(row=checkout_row, column=col_num, value=guest[column_name])
            checkout_row += 1
        elif guest['action'] in ('URGENT - CLASH', 'Property not allocated'):
            # Add the "ERRORS" header and column names
            if errors_row == 2:
                sheet.cell(row=1, column=1).font = Font(bold=True)
                sheet.cell(row=1, column=1, value='ERRORS')
                for col_num, column_name in enumerate(column_names, start=2):
                    sheet.cell(row=1, column=col_num).font = Font(bold=True)
                    sheet.cell(row=1, column=col_num, value=column_name)
            # Add the data for "URGENT - CLASH" rows
            for col_num, column_name in enumerate(column_names, start=1):
                sheet.cell(row=errors_row, column=col_num, value=guest[column_name])
            errors_row += 1


    # Save the Excel file
    workbook.save("organized_data.xlsx")
def set_column_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

def send_email(action_list):
    #organize_data(action_list)
    # Create Excel workbook
    workbook = openpyxl.Workbook()

    # Create action sheet
    action_sheet = workbook.active
    action_sheet.title = "Action"
    action_sheet.append(['guest', 'tag', 'allocated_property', 'flat_booked', 'checkin_date', 'checkout_date', 'checkin_time', 'checkout_time', 'group', 'action', 'keys'])
    for guest in action_list:
        if "action" not in guest:
            guest.update({"action": ""})
        
        if "keys" not in guest:
            guest.update({"keys": ""})
        
        action_sheet.append([guest['guest'], guest['tag'], guest['allocated_property'], guest['flat_booked'], guest['checkin_date'], guest['checkout_date'], guest['checkin_time'], guest['checkout_time'], guest['group'], guest['action'], guest['keys']])
    
    set_column_width(action_sheet)

    # Save the Excel file
    workbook.save("organized_data.xlsx")
    workbook.close()

    formatExcel()

    # Save the Excel file
    filename = "output.xlsx"
    attachment_path = os.path.abspath(filename)
    #workbook.save(filename)
    
    
    
    # Email details
    sender_email = 'smtp@theflexliving.com'
    receiver_emails = ['sid@theflexliving.com', 'michael@theflexliving.com', 'info@theflexliving.com']
    subject = "[AUTOMATION] - Check-in Check-out of today"
    body = "Please find attached the Excel file containing the list of next check-ins and check-outs"

    # Create message container
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ", ".join(receiver_emails)
    msg['Subject'] = subject

    # Add body to email
    msg.attach(MIMEText(body, "plain"))

    # Attach the Excel file
    with open(attachment_path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read()) 
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {filename}",
        )
        msg.attach(part)

    # Convert message object to string
    text = msg.as_string()

    # Log in to SMTP server and send email
    try:
        smtp_server = smtplib.SMTP('smtp.gmail.com', 587)
        smtp_server.starttls()
        smtp_server.login(sender_email, "")
        smtp_server.sendmail(sender_email, receiver_emails, text)
        smtp_server.quit()
    
    except Exception as e:
        print("An error occurred while sending the email:", e)

def surligner_actions(csv_file):
    # Ouvrir le fichier CSV en lecture
    with open(csv_file, 'r') as file:
        # Lire le contenu du fichier CSV
        csv_data = csv.reader(file)
        header = next(csv_data)  # Lire l'en-tête du fichier CSV

        # Trouver l'indice de la colonne "Action" dans l'en-tête
        action_index = header.index('action')

        # Créer une liste pour stocker les lignes du fichier CSV modifié
        modified_rows = []

        # Parcourir les lignes du fichier CSV
        for row in csv_data:
            action = row[action_index].lower()

            # Vérifier le type d'action et ajouter le code de couleur correspondant
            if 'SEND CHECKOUT' in action:
                row[action_index] = f"\033[91m{row[action_index]}\033[0m"  # Rouge
            elif 'NO CLASH' in action:
                row[action_index] = f"\033[92m{row[action_index]}\033[0m"  # Vert
            elif 'RELOCATE' in action:
                row[action_index] = f"\033[93m{row[action_index]}\033[0m"  # Orange

            modified_rows.append(row)  # Ajouter la ligne modifiée à la liste

    # Écrire les lignes modifiées dans un nouveau fichier CSV
    with open('modified_actions.csv', 'w', newline='') as file:
        csv_writer = csv.writer(file)
        csv_writer.writerow(header)  # Écrire l'en-tête dans le nouveau fichier CSV
        csv_writer.writerows(modified_rows)  # Écrire les lignes modifiées

    #print("Le fichier CSV a été modifié avec succès.")

def main():
    today_date = date.today()
    today = today_date.strftime("%d/%m/%Y")
    
    items = get_formated_items()
    keynest = KeynestClient()
    keynest_keys_json = keynest.keynest_client.get_key_nests('keys/')
    if keynest_keys_json:
        keys_list = keynest_keys_json['ResponsePacket']['KeyList']
    
    for key in keys_list :
        #new_key=key['KeyName'].split('(')[0].strip()
        key.update({"KeyName" : key['KeyName'].split('(')[0].strip()})


    # Check overlapping booking
    # 1 / Filter properties
    for item in items:
        filtered_properties = [prop for prop in items if prop['allocated_property'] == item['allocated_property']]

        not_allocated_properties = [prop for prop in items if prop['group'] == "1 Week Before Check In" and (prop['allocated_property'] == None or prop['allocated_property'] == "" )]

        # Today date
        today = datetime.now().date()

        # Convertir la date donnée en paramètre en objet de type 'date'
        checkin_date = datetime.strptime(item['checkin_date'], "%d/%m/%Y").date()
        checkout_date = datetime.strptime(item['checkout_date'], "%d/%m/%Y").date()

        # Calculer la différence entre les deux dates
        checkin_diff = (checkin_date - today).days
        checkout_diff = (checkout_date - today).days

        if 0 <= checkin_diff <= 7 and item['group'] != "Current tenants":

            # Not allocated properties in 1 Week Before Check In
            if item in not_allocated_properties:
                action=item
                print(f"{item['guest']} - ERROR")
                action.update({"action" : "Property not allocated"})
                action_list.append(action)

        if item['allocated_property'] != "" and item['allocated_property'] != None and item['flat_booked'] != "" and item['flat_booked'] != None and item['status'] != "Cancelled" and item['allocated_property'].lower() != "to be cancelled":
            # Send checkin checkout

            
            
            if 0 <= checkin_diff <= 7 and item['group'] != "Current tenants":

                # Relocation : if flat booked and allocated property are different and tag is different as well 
                flat_booked_tag = item['flat_booked'].split(' - ')[0]
                
                if item['allocated_property'] != item['flat_booked'] and item['group'] != "Current tenants" and item['group'] != "Check Outs" and item['tag'] != flat_booked_tag :
                    action= item
                    action.update({"action": "RELOCATE"})

                # Delay Checkin 
                if item['allocated_property'] == item['flat_booked'] and item['group'] != "Current tenants":                
                    
                    if len(filtered_properties) > 1 :

                        # Compare check-in 1st property & check-in 2nd property 
                        check_in_1=datetime.strptime(filtered_properties[0]['checkin_date'],"%d/%m/%Y").date()
                        check_in_2=datetime.strptime(filtered_properties[1]['checkin_date'],"%d/%m/%Y").date()
                        
                        if check_in_1 > check_in_2 :
                        
                            # Define before and after item / reservation
                            before = filtered_properties[1]
                            after = filtered_properties[0]

                            # Convertir la date donnée en paramètre en objet de type 'date'
                            check_in_before = datetime.strptime(before['checkin_date'], "%d/%m/%Y").date()
                            check_out_before = datetime.strptime(before['checkout_date'], "%d/%m/%Y").date()
                            check_in_after = datetime.strptime(after['checkin_date'], "%d/%m/%Y").date()
                            check_out_after = datetime.strptime(after['checkout_date'], "%d/%m/%Y").date()
                            
                            # Calculer la différence de jours
                            diff=(check_out_before - check_in_after).days

                        else:   
                            
                            # Define before and after item / reservation
                            before = filtered_properties[0]
                            after = filtered_properties[1]
                            
                            # Convertir la date donnée en paramètre en objet de type 'date'
                            check_in_before = datetime.strptime(before['checkin_date'], "%d/%m/%Y").date()
                            check_out_before = datetime.strptime(before['checkout_date'], "%d/%m/%Y").date()
                            check_in_after = datetime.strptime(after['checkin_date'], "%d/%m/%Y").date()
                            check_out_after = datetime.strptime(after['checkout_date'], "%d/%m/%Y").date()
                            
                            # Calculer la différence de jours
                            diff=(check_out_before - check_in_after).days

                        if diff > 0 :

                            action= after
                            action.update({"action": "URGENT - CLASH"})
                            
                            booking_length = (check_out_after - check_in_after).days
                            delay=(check_in_after - check_out_before).days

                            if booking_length > 9 and delay < 6 :
                                action.update({"action": "URGENT - CLASH - Ask for delayed Checking"})
                                
                            elif booking_length < 10 and delay < 4 :
                                action.update({"action": "URGENT - CLASH - Ask for delayed Checking"})
                                
                            else : 
                                action.update({"action": "URGENT - CLASH - No delayed checkin possible"})
                        
                        if diff == 0 :
                            # Define values
                            default_check_in_time = datetime.strptime("03:00 PM", "%I:%M %p")
                            default_check_out_time = datetime.strptime("10:00 AM", "%I:%M %p")
                            if before['checkout_time'] != None and before['checkout_time'] != "" and after['checkin_time'] != None and after['checkin_time'] != "" :
                                check_out_time = datetime.strptime(before['checkout_time'], "%I:%M %p")
                                check_in_time = datetime.strptime(after['checkin_time'], "%I:%M %p")
                            
                                # Compare
                                if check_in_time > default_check_in_time or check_out_time < default_check_out_time :
                                    #print("EARLY CHECK-IN & LATE CHECK-OUT")
                                    action= after
                                    action.update({"action": "URGENT - Early check-in & Late check-out "})

                if "action" not in item:
                    action=item
                    if action['form'] == "Completed":
                        action.update({"action" : "SEND CHECKIN INSTRUCTIONS - CHECK-IN FORM COMPLETED"})
                    else:
                        action.update({"action" : "CHECK-IN FORM NOT COMPLETED - SEND REMINDER"})
                
                action= item
                action_list.append(action)

            if checkin_diff == -1 :
                no_show_keys(item, keys_list)
                action= item
                action_list.append(action)

            if 0 <= checkout_diff <= 3:      
                # Check if keys are in store
                if checkout_diff == 0 :
                    action=keys_dropped(item, keys_list)
                else :    
                    action= item
                    action.update({"action": "SEND CHECKOUT INSTRUCTIONS"})
            
                action_list.append(action)
            
    new_action_list = []
    for item in action_list:
        if item not in new_action_list:
            new_action_list.append(item)
    send_email(new_action_list)

    print("*** SCRIPT ENDED WITH NO ERROR *** ")   

    

if __name__ == "__main__":
    main() 